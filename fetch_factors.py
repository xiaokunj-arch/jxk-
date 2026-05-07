# -*- coding: utf-8 -*-
"""
从免费数据源自动下载宏观/价格因子，写入 大宗商品轮动_数据2.xlsx
数据来源：
  FRED   - 实际利率 / 美元指数
  EIA    - 美国商业原油库存（解析官网 HTML）
  FRED   - 美国工业产出指数（作为 PMI 替代，PMI 需从 Wind 手动补充）
  yfinance - 美元指数DXY / 黄金ETF价格 / 各品种期货价格
"""

import requests
import pandas as pd
import openpyxl
import yfinance as yf
from io import StringIO
from datetime import datetime
import time

try:
    import akshare as ak
    HAS_AKSHARE = True
except ImportError:
    HAS_AKSHARE = False

MASTER_FILE = "大宗商品轮动_数据2.xlsx"
START = "1990-01-01"


# ══════════════════════════════════════════
# 1. FRED 下载
# ══════════════════════════════════════════
FRED_CONFIGS = [
    # (series_id,  子表名,        指标中文名)
    ("DFII10",   "实际利率",   "美国10年期TIPS收益率(实际利率,%)"),
    ("DTWEXBGS", "美元指数_FRED", "美元指数(美联储广义美元指数)"),
    ("INDPRO",   "工业产出指数", "美国工业生产指数(2017=100)"),   # PMI 替代指标
]


def download_fred(series_id: str, col_name: str) -> pd.DataFrame | None:
    url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}"
    try:
        # 优先用系统 curl（requests 对 FRED 有时超时）
        import subprocess
        result = subprocess.run(
            ["curl", "-s", "--max-time", "30", url],
            capture_output=True, text=True
        )
        text = result.stdout if result.returncode == 0 and result.stdout.startswith("observation") else None
        if text is None:
            resp = requests.get(url, timeout=30, headers={"User-Agent": "Mozilla/5.0"})
            resp.raise_for_status()
            text = resp.text
        df = pd.read_csv(StringIO(text))
        df.columns = ["日期", col_name]
        df["日期"] = pd.to_datetime(df["日期"])
        df = df[df["日期"] >= START].copy()
        df[col_name] = pd.to_numeric(df[col_name], errors="coerce")
        return df.dropna().sort_values("日期").reset_index(drop=True)
    except Exception as e:
        print(f"    ❌ FRED {series_id} 失败: {e}")
        return None


# ══════════════════════════════════════════
# 2. EIA 原油库存
# ══════════════════════════════════════════
def download_eia_crude_inventory() -> pd.DataFrame | None:
    """解析 EIA 官网 HTML，获取美国商业原油周度库存"""
    url = "https://www.eia.gov/dnav/pet/hist/LeafHandler.ashx?n=PET&s=WCRSTUS1&f=W"
    try:
        r = requests.get(url, timeout=60, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        tables = pd.read_html(StringIO(r.text))
        # 取形状最大的表格（宽格式，按 年月×周）
        raw = max(tables, key=lambda t: t.shape[0] * t.shape[1])
        # 列结构：Year-Month | Week1 EndDate | Week1 Value | Week2 EndDate | Week2 Value | ...
        raw.columns = [f"col{i}" for i in range(raw.shape[1])]
        raw = raw.iloc[2:].reset_index(drop=True)   # 跳过多级表头2行

        rows = []
        for _, r2 in raw.iterrows():
            year_month = str(r2["col0"]).strip()
            if not year_month or year_month in ("nan", "NaN"):
                continue
            # 每周数据占 2 列：EndDate + Value，从 col1 开始
            for j in range(1, raw.shape[1] - 1, 2):
                end_date_raw = str(r2[f"col{j}"]).strip()
                value_raw = r2[f"col{j+1}"]
                if end_date_raw in ("nan", "NaN", "") or pd.isna(value_raw):
                    continue
                try:
                    # end_date_raw 格式 MM/DD，year_month 格式 YYYY-MMM(英) 或 YYYY-Mon
                    year = year_month.split("-")[0]
                    month_day = end_date_raw  # "03/07"
                    d = datetime.strptime(f"{year}/{month_day}", "%Y/%m/%d").date()
                    rows.append((d, float(value_raw)))
                except Exception:
                    continue

        df = pd.DataFrame(rows, columns=["日期", "美国商业原油库存(千桶)"])
        df["日期"] = pd.to_datetime(df["日期"])
        df = df[df["日期"] >= START].drop_duplicates("日期").sort_values("日期").reset_index(drop=True)
        return df
    except Exception as e:
        print(f"    ❌ EIA 原油库存失败: {e}")
        return None


# ══════════════════════════════════════════
# 3. yfinance 价格数据
# ══════════════════════════════════════════
YFINANCE_CONFIGS = [
    # (ticker,       子表名,             指标中文名,                           field)
    ("DX-Y.NYB",  "美元指数DXY",       "美元指数DXY(ICE)",                   "Close"),
    ("^VIX",      "VIX恐慌指数",        "CBOE波动率指数(VIX)",                "Close"),
    ("FXI",       "FXI中国大盘ETF",     "iShares中国大盘ETF收盘价(USD)",       "Close"),
    ("TTF=F",     "TTF欧洲天然气",      "TTF欧洲天然气期货收盘价(EUR/MWh)",    "Close"),
    ("GLD",       "黄金ETF价格",        "SPDR黄金ETF收盘价(USD)",              "Close"),
    ("GC=F",      "黄金期货_YF",        "COMEX黄金期货收盘价(USD/oz)",         "Close"),
    ("SI=F",      "银期货_YF",          "COMEX银期货收盘价(USD/oz)",           "Close"),
    ("HG=F",      "铜期货_YF",          "COMEX铜期货收盘价(USD/lb)",           "Close"),
    ("CL=F",      "WTI原油期货_YF",     "WTI原油期货收盘价(USD/bbl)",          "Close"),
    ("BZ=F",      "布伦特原油期货_YF",  "布伦特原油期货收盘价(USD/bbl)",       "Close"),
]


def download_yfinance(ticker: str, col_name: str, field: str = "Close") -> pd.DataFrame | None:
    try:
        raw = yf.download(ticker, start=START, auto_adjust=True, progress=False)
        if raw.empty:
            print(f"    ❌ {ticker} 无数据")
            return None
        df = raw[[field]].copy()
        # yfinance 1.x MultiIndex columns 兼容
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = [col_name]
        else:
            df.columns = [col_name]
        df.index.name = "日期"
        df = df.reset_index()
        df["日期"] = pd.to_datetime(df["日期"])
        df = df.dropna().sort_values("日期").reset_index(drop=True)
        return df
    except Exception as e:
        print(f"    ❌ yfinance {ticker} 失败: {e}")
        return None


# ══════════════════════════════════════════
# 4. 写入 Excel
# ══════════════════════════════════════════
def write_sheet(wb: openpyxl.Workbook, sheet_name: str, df: pd.DataFrame):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    headers = list(df.columns)
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    for r, row in enumerate(df.itertuples(index=False), start=2):
        for c, val in enumerate(row, 1):
            if hasattr(val, "date"):
                ws.cell(r, c, val.date())
            elif pd.isna(val):
                ws.cell(r, c, None)
            else:
                ws.cell(r, c, float(val) if isinstance(val, (int, float)) else val)


# ══════════════════════════════════════════
# 主程序
# ══════════════════════════════════════════
def main():
    if not __import__("os").path.exists(MASTER_FILE):
        print(f"❌ 找不到主文件 {MASTER_FILE}，请先运行 generate_new_excel.py")
        return

    wb = openpyxl.load_workbook(MASTER_FILE)
    results, failed = [], []

    # ── FRED ──
    print("\n【FRED 数据】")
    for series_id, sheet_name, col_name in FRED_CONFIGS:
        print(f"  ↓ {sheet_name} ({series_id})...", end=" ", flush=True)
        df = download_fred(series_id, col_name)
        if df is not None and not df.empty:
            write_sheet(wb, sheet_name, df)
            print(f"✅ {len(df)} 行")
            results.append(sheet_name)
        else:
            failed.append(sheet_name)
        time.sleep(0.8)

    # ── EIA 原油库存 ──
    print("\n【EIA 原油库存】")
    print("  ↓ 原油库存...", end=" ", flush=True)
    df_oil = download_eia_crude_inventory()
    if df_oil is not None and not df_oil.empty:
        write_sheet(wb, "原油库存", df_oil)
        print(f"✅ {len(df_oil)} 行（周度数据）")
        results.append("原油库存")
    else:
        failed.append("原油库存")

    # ── yfinance ──
    print("\n【yfinance 价格数据】")
    for ticker, sheet_name, col_name, field in YFINANCE_CONFIGS:
        print(f"  ↓ {sheet_name} ({ticker})...", end=" ", flush=True)
        df = download_yfinance(ticker, col_name, field)
        if df is not None and not df.empty:
            write_sheet(wb, sheet_name, df)
            print(f"✅ {len(df)} 行")
            results.append(sheet_name)
        else:
            failed.append(sheet_name)
        time.sleep(0.3)

    # ── 美债收益率（FRED：10年期、2年期名义收益率 + 10Y-2Y利差）──
    print("\n【美债收益率】")
    try:
        treasury_series = [
            ("DGS10", "美国10年期国债收益率(%)"),
            ("DGS2",  "美国2年期国债收益率(%)"),
            ("T10Y2Y","美国10Y-2Y利差(%)"),
        ]
        dfs_t = []
        for sid, col in treasury_series:
            df_t = download_fred(sid, col)
            if df_t is not None and not df_t.empty:
                dfs_t.append(df_t.set_index("日期"))
        if dfs_t:
            merged_t = pd.concat(dfs_t, axis=1).reset_index().sort_values("日期")
            write_sheet(wb, "美债收益率", merged_t)
            print(f"✅ {len(merged_t)} 行，最新: {merged_t['日期'].iloc[-1].date()}")
            results.append("美债收益率")
        else:
            print("❌ 所有序列均失败")
            failed.append("美债收益率")
    except Exception as e:
        print(f"❌ {e}")
        failed.append("美债收益率")

    # ── 中国制造业PMI + 中国PPI（akshare）──
    if HAS_AKSHARE:
        print("\n【中国宏观数据（akshare）】")

        # 中国制造业PMI
        print("  ↓ 中国制造业PMI...", end=" ", flush=True)
        try:
            df_pmi = ak.macro_china_pmi_yearly()
            df_pmi = df_pmi[["日期", "今值"]].copy()
            df_pmi.columns = ["日期", "中国官方制造业PMI"]
            df_pmi["日期"] = pd.to_datetime(df_pmi["日期"])
            df_pmi["中国官方制造业PMI"] = pd.to_numeric(df_pmi["中国官方制造业PMI"], errors="coerce")
            df_pmi = df_pmi.dropna().sort_values("日期").reset_index(drop=True)
            write_sheet(wb, "中国制造业PMI", df_pmi)
            print(f"✅ {len(df_pmi)} 行，最新: {df_pmi['日期'].iloc[-1].date()}")
            results.append("中国制造业PMI")
        except Exception as e:
            print(f"❌ {e}")
            failed.append("中国制造业PMI")
        time.sleep(0.5)

        # 中国PPI
        print("  ↓ 中国PPI...", end=" ", flush=True)
        try:
            df_ppi = ak.macro_china_ppi_yearly()
            df_ppi = df_ppi[["日期", "今值"]].copy()
            df_ppi.columns = ["日期", "中国PPI同比(%)"]
            df_ppi["日期"] = pd.to_datetime(df_ppi["日期"])
            df_ppi["中国PPI同比(%)"] = pd.to_numeric(df_ppi["中国PPI同比(%)"], errors="coerce")
            df_ppi = df_ppi.dropna().sort_values("日期").reset_index(drop=True)
            write_sheet(wb, "中国ppi", df_ppi)
            print(f"✅ {len(df_ppi)} 行，最新: {df_ppi['日期'].iloc[-1].date()}")
            results.append("中国ppi")
        except Exception as e:
            print(f"❌ {e}")
            failed.append("中国ppi")
        time.sleep(0.5)
    else:
        print("\n⚠️ akshare 未安装，跳过中国PMI/PPI自动下载（可从 Wind 手动补充）")

    # ── 焦煤期货（大商所JM主力合约，替代已停牌的郑商所ZC动力煤）──
    print("\n【焦煤期货价格（大商所JM0）】")
    print("  ↓ 焦煤主力合约（期货价格表第9/10列）...", end=" ", flush=True)
    try:
        df_jm = ak.futures_zh_daily_sina(symbol="JM0")
        df_jm["date"] = pd.to_datetime(df_jm["date"])
        df_jm = df_jm.sort_values("date").dropna(subset=["close"])

        ws_px = wb["期货价格"]
        ws_px.cell(1, 9, "日期")
        ws_px.cell(1, 10, "期货收盘价(连续):大商所焦煤JM")
        for row in ws_px.iter_rows(min_row=2, min_col=9, max_col=10):
            for cell in row:
                cell.value = None
        for i, (_, row) in enumerate(df_jm.iterrows(), start=2):
            ws_px.cell(i, 9, row["date"].date())
            ws_px.cell(i, 10, float(row["close"]))
        print(f"✅ {len(df_jm)} 行，最新: {df_jm['date'].iloc[-1].date()}")
        results.append("焦煤期货(JM0)")
    except Exception as e:
        print(f"❌ {e}")
        failed.append("焦煤期货(JM0)")

    # ── 中国信贷脉冲（M2 + GDP 计算） ──
    print("\n【中国信贷脉冲】")
    try:
        df_m2 = ak.macro_china_supply_of_money()
        df_m2 = df_m2[["统计时间", "货币和准货币（广义货币M2）"]].copy()
        df_m2.columns = ["date_str", "m2"]
        df_m2["m2"] = pd.to_numeric(df_m2["m2"], errors="coerce")
        def _parse_m2_date(s):
            p = str(s).split(".")
            if len(p) == 2:
                return pd.Timestamp(year=int(p[0]), month=int(p[1]), day=1) + pd.offsets.MonthEnd(0)
            return pd.NaT
        df_m2["date"] = df_m2["date_str"].apply(_parse_m2_date)
        m2 = df_m2.dropna(subset=["date"]).set_index("date").sort_index()[["m2"]]
        m2["m2_new_12m"] = m2["m2"] - m2["m2"].shift(12)

        df_gdp_raw = ak.macro_china_gdp()
        single_q = df_gdp_raw[df_gdp_raw["季度"].str.match(r"^\d{4}年第[1-4]季度$")].copy()
        def _parse_quarter(s):
            year, q = int(s[:4]), int(s[6])
            return pd.Timestamp(year=year, month=q * 3, day=1) + pd.offsets.MonthEnd(0)
        single_q["date"] = single_q["季度"].apply(_parse_quarter)
        gdp = single_q[["date", "国内生产总值-绝对值"]].rename(columns={"国内生产总值-绝对值": "gdp_q"})
        gdp = gdp.sort_values("date").set_index("date").dropna()
        gdp["gdp_annual"] = gdp["gdp_q"].rolling(4, min_periods=4).sum()
        gdp_monthly = gdp[["gdp_annual"]].resample("ME").last().ffill()

        merged = m2.join(gdp_monthly, how="left")
        merged["gdp_annual"] = merged["gdp_annual"].ffill()
        merged["delta_m2_new"] = merged["m2_new_12m"] - merged["m2_new_12m"].shift(12)
        merged["credit_impulse"] = merged["delta_m2_new"] / merged["gdp_annual"] * 100
        merged["m2_yoy"] = (m2["m2"] / m2["m2"].shift(12) - 1) * 100

        result = merged[["m2", "m2_new_12m", "m2_yoy", "credit_impulse"]].dropna(subset=["credit_impulse"]).reset_index()
        result.columns = ["日期", "M2余额(亿元)", "12个月新增M2(亿元)", "M2同比增速(%)", "中国信贷脉冲(%)"]

        write_sheet(wb, "中国信贷脉冲", result)
        print(f"✅ {len(result)} 行，最新: {result['日期'].iloc[-1].date()}")
        results.append("中国信贷脉冲")
    except Exception as e:
        print(f"❌ {e}")
        failed.append("中国信贷脉冲")

    wb.save(MASTER_FILE)

    print("\n" + "═" * 55)
    print(f"✅ 成功写入 {len(results)} 个子表：{results}")
    if failed:
        print(f"❌ 未能下载：{failed}")
    print("""
【说明】
• 实际利率       → FRED DFII10（10年期TIPS收益率）
• 美元指数       → FRED广义美元指数 + ICE DXY（两个版本）
• 原油库存       → EIA官网，周度，美国商业原油库存（千桶）
• 工业产出指数   → FRED INDPRO，作为 PMI 替代指标
• 各品种价格     → Yahoo Finance（期货近月合约）
• GLD价格        → SPDR黄金ETF收盘价（可作为黄金 ETF 持仓的代理）

【还需要从 Wind 手动补充的数据】
• 美国ISM制造业PMI  → Wind 搜索"美国ISM制造业PMI"或代码 M0000545
• LME铜库存        → Wind 搜索"LME铜库存"
• COMEX银库存      → Wind 搜索"COMEX银库存"
• 黄金ETF实物持仓（吨数）→ Wind 搜索"SPDR黄金持仓"
""")


if __name__ == "__main__":
    main()
