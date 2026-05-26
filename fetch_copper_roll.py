# -*- coding: utf-8 -*-
"""
计算并保存 SHFE 铜期货展期收益率（年化）到主数据文件。

原理：
  每月15日（就近取工作日）从上期所日结算数据中取：
    近月合约（排序第一）结算价  → front
    当日日期 +3 个自然月对应合约 → defer
  展期收益率 = (front / defer)^4 - 1   （年化，基于3个月周期）

存储：
  写入 "大宗商品轮动_数据2.xlsx" → sheet "铜展期收益"
  格式：日期（升序）| 铜展期收益率

首次运行约 2-3 分钟（216 次 API 调用，4 线程并发）；
增量运行只补充缺失月份，通常 <10 秒。
"""
from __future__ import annotations
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, timedelta
from pathlib import Path

import akshare as ak
import openpyxl
import pandas as pd

MASTER_FILE = "大宗商品轮动_数据2.xlsx"
SHEET_NAME = "铜展期收益"
START_YEAR = 2008
MAX_WORKERS = 4


# ── 生成采样日期：每月约第15日最近的工作日（周一~周五）──────────────
def _generate_sample_dates(from_date: date, to_date: date) -> list[date]:
    """每月生成一个采样日期（15日起往后找，最多±5天内的工作日）。"""
    results = []
    y, m = from_date.year, from_date.month
    while date(y, m, 1) <= to_date:
        # 从每月15日开始，找最近工作日
        for delta in range(0, 6):
            d = date(y, m, min(15 + delta, 28))
            if d.weekday() < 5 and d <= to_date:  # 周一~周五
                results.append(d)
                break
        else:
            # fallback: 向前找
            for delta in range(1, 6):
                d = date(y, m, min(15 - delta, 28))
                if d.weekday() < 5:
                    results.append(d)
                    break
        # 下个月
        m += 1
        if m > 12:
            m = 1
            y += 1
    return results


# ── 单次 API 调用 ────────────────────────────────────────────────────
def _fetch_one(d: date) -> tuple[date, float | None]:
    """
    获取日期 d 的 SHFE 铜展期收益率。
    远月合约 = d + 3个自然月对应的 CU 合约。
    返回 (date, roll_yield) 或 (date, None)。
    """
    date_str = d.strftime("%Y%m%d")
    try:
        df = ak.get_shfe_daily(date=date_str)
        cu = df[df["symbol"].str.startswith("CU")].sort_values("symbol").reset_index(drop=True)
        if cu.empty:
            return d, None

        front = float(cu.iloc[0]["settle"])

        # 计算 +3 个月的合约代码
        target_month = d.month + 3
        target_year = d.year
        if target_month > 12:
            target_month -= 12
            target_year += 1
        target_symbol = f"CU{target_year % 100:02d}{target_month:02d}"

        row = cu[cu["symbol"] == target_symbol]
        if row.empty:
            # 备用：直接用 index[3]（约3-4个月后到期）
            if len(cu) >= 4:
                row = cu.iloc[[3]]
            else:
                return d, None

        defer = float(row.iloc[0]["settle"])
        if front <= 0 or defer <= 0:
            return d, None

        roll_yield = (front / defer) ** 4 - 1  # 年化（3个月期）
        return d, round(roll_yield, 6)
    except Exception:
        return d, None


# ── 获取现有数据的最新日期 ─────────────────────────────────────────
def _get_existing_latest(wb: openpyxl.Workbook) -> date | None:
    if SHEET_NAME not in wb.sheetnames:
        return None
    ws = wb[SHEET_NAME]
    last_dt = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            try:
                last_dt = pd.Timestamp(row[0]).date()
            except Exception:
                pass
    return last_dt


# ── 写入到 sheet ──────────────────────────────────────────────────
def _write_sheet(wb: openpyxl.Workbook, df: pd.DataFrame):
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)
    ws.cell(1, 1, "日期")
    ws.cell(1, 2, "铜展期收益率")
    for i, (dt, val) in enumerate(df.itertuples(), start=2):
        ws.cell(i, 1, dt if isinstance(dt, date) else dt.date())
        ws.cell(i, 2, float(val))


# ── 主流程 ────────────────────────────────────────────────────────
def main():
    wb = openpyxl.load_workbook(MASTER_FILE)
    existing_latest = _get_existing_latest(wb)
    print(f"  现有最新数据: {existing_latest or '无'}")

    from_date = date(START_YEAR, 1, 1) if existing_latest is None else (
        date(existing_latest.year, existing_latest.month, 1) + timedelta(days=32)
    )
    from_date = date(from_date.year, from_date.month, 1)
    to_date = date.today()

    if from_date > to_date:
        print("  ✅ 已是最新，无需更新。")
        wb.close()
        return

    sample_dates = _generate_sample_dates(from_date, to_date)
    print(f"  需下载 {len(sample_dates)} 个月份数据（{from_date} → {to_date}）")
    print(f"  使用 {MAX_WORKERS} 线程并发，预计 {len(sample_dates) * 2.3 / MAX_WORKERS / 60:.1f} 分钟...")

    # 多线程下载
    fetch_results: dict[date, float] = {}
    lock = threading.Lock()
    completed = [0]

    def _worker(d: date):
        dt, val = _fetch_one(d)
        with lock:
            completed[0] += 1
            if val is not None:
                fetch_results[dt] = val
            if completed[0] % 20 == 0 or completed[0] == len(sample_dates):
                print(f"    [{completed[0]}/{len(sample_dates)}] 已完成，有效数据 {len(fetch_results)} 条")

    t0 = time.time()
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        list(executor.map(_worker, sample_dates))

    elapsed = time.time() - t0
    print(f"  下载完成，用时 {elapsed:.0f}s，获取 {len(fetch_results)} 条有效记录")

    if not fetch_results:
        print("  ❌ 无有效数据，跳过写入。")
        wb.close()
        return

    # 合并旧数据
    new_df = pd.DataFrame(
        [(dt, val) for dt, val in sorted(fetch_results.items())],
        columns=["日期", "铜展期收益率"],
    ).set_index("日期")

    if existing_latest is not None and SHEET_NAME in wb.sheetnames:
        old_ws = wb[SHEET_NAME]
        old_rows = [(row[0], row[1]) for row in old_ws.iter_rows(min_row=2, values_only=True)
                    if row and row[0] is not None and row[1] is not None]
        if old_rows:
            old_df = pd.DataFrame(old_rows, columns=["日期", "铜展期收益率"])
            old_df["日期"] = pd.to_datetime(old_df["日期"]).dt.date
            old_df = old_df.set_index("日期")
            new_df = pd.concat([old_df, new_df]).sort_index()
            new_df = new_df[~new_df.index.duplicated(keep="last")]

    _write_sheet(wb, new_df)
    wb.save(MASTER_FILE)
    print(f"  ✅ 已写入 {len(new_df)} 行 → sheet「{SHEET_NAME}」")
    print(f"  范围：{new_df.index[0]} → {new_df.index[-1]}")
    print(f"  最近5条:\n{new_df.tail()}")


if __name__ == "__main__":
    main()
