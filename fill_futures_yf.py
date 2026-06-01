# -*- coding: utf-8 -*-
"""
用 yfinance 补充「期货价格」sheet 中黄金/白银/铜/原油的最新价格数据。
仅追加缺失日期，不覆盖已有 Wind 数据。
"""

import openpyxl
import yfinance as yf
import pandas as pd
from datetime import date

MASTER_FILE = "大宗商品轮动_数据2.xlsx"

# 期货价格 sheet 中各品种的（日期列索引, 价格列索引）和对应的 yfinance ticker
ASSETS = [
    ("黄金",       0, 1, "GC=F"),
    ("白银",       2, 3, "SI=F"),
    ("大成有色ETF", 4, 5, "512400.SS"),  # 铜→大成有色ETF，Wind手动维护原油LOF(160416)和煤炭LOF(160819)
]

START = "2005-01-01"


def main():
    wb = openpyxl.load_workbook(MASTER_FILE)
    ws = wb["期货价格"]
    rows = list(ws.iter_rows(values_only=True))

    for name, dc, pc, ticker in ASSETS:
        # 找该品种在 sheet 中已有的最新日期
        existing_dates = set()
        last_row_idx = 1  # 1-based，表头在第1行
        for i, row in enumerate(rows[1:], start=2):
            if row[dc] is not None:
                try:
                    d = pd.Timestamp(row[dc]).date()
                    existing_dates.add(d)
                    last_row_idx = i
                except Exception:
                    pass

        max_date = max(existing_dates) if existing_dates else date(2005, 1, 1)
        print(f"  {name}: 已有数据到 {max_date}，ticker={ticker}")

        # 下载 yfinance 数据
        df = yf.download(ticker, start=str(max_date), auto_adjust=True, progress=False)
        if df.empty:
            print(f"    ❌ {ticker} 无数据")
            continue

        close = df["Close"]
        if isinstance(close, pd.DataFrame):
            close = close.iloc[:, 0]
        close = close.dropna()
        close.index = pd.to_datetime(close.index).normalize()

        # 只保留比现有数据更新的日期
        today = pd.Timestamp.today().date()
        new_rows = [(d.date(), float(v)) for d, v in close.items()
                    if max_date < pd.Timestamp(d).date() < today]

        if not new_rows:
            print(f"    ✅ {name} 已是最新，无需追加")
            continue

        # 追加新行到对应列（从 last_row_idx+1 开始）
        for offset, (d, v) in enumerate(new_rows):
            row_num = last_row_idx + 1 + offset
            ws.cell(row=row_num, column=dc + 1, value=d)
            ws.cell(row=row_num, column=pc + 1, value=v)

        print(f"    ✅ {name} 追加 {len(new_rows)} 行，最新: {new_rows[-1][0]}")

    wb.save(MASTER_FILE)
    print("\n期货价格 sheet 已更新保存。")


if __name__ == "__main__":
    main()
