# -*- coding: utf-8 -*-
"""
下载 CFTC COT 持仓报告（Disaggregated Futures），提取管理资金净头寸。

用法：
    python3 fetch_cot.py

输出：在大宗商品轮动_数据2.xlsx 中写入以下 sheet：
    - COT黄金    管理资金净头寸（多 - 空）
    - COT白银
    - COT铜
    - COT原油
"""

from __future__ import annotations

import io
import zipfile
from pathlib import Path

import pandas as pd
import requests

MASTER_FILE = Path("大宗商品轮动_数据2.xlsx")

# CFTC Disaggregated Futures Report 年度文件 URL
URL_TEMPLATE = "https://www.cftc.gov/files/dea/history/com_disagg_txt_{year}.zip"
URL_HIST = "https://www.cftc.gov/files/dea/history/com_disagg_txt_hist_2006_2016.zip"

# 合约名称（历年有轻微改动，全部列出做模糊匹配）
CONTRACT_PATTERNS = {
    "gold_cot":   ["GOLD - COMMODITY EXCHANGE INC."],
    "silver_cot": ["SILVER - COMMODITY EXCHANGE INC."],
    "copper_cot": ["COPPER-GRADE #1 - COMMODITY EXCHANGE INC.",
                   "COPPER- #1 - COMMODITY EXCHANGE INC."],
    "oil_cot":    ["CRUDE OIL, LIGHT SWEET - NEW YORK MERCANTILE EXCHANGE",
                   "WTI-PHYSICAL - NEW YORK MERCANTILE EXCHANGE"],
}

COT_SHEET_NAMES = {
    "gold_cot":   "COT黄金",
    "silver_cot": "COT白银",
    "copper_cot": "COT铜",
    "oil_cot":    "COT原油",
}

COT_DESCRIPTIONS = {
    "gold_cot":   "COMEX黄金期货管理资金净头寸（多-空，手）",
    "silver_cot": "COMEX白银期货管理资金净头寸（多-空，手）",
    "copper_cot": "COMEX铜期货管理资金净头寸（多-空，手）",
    "oil_cot":    "NYMEX原油期货管理资金净头寸（多-空，手）",
}


def fetch_zip(url: str) -> pd.DataFrame | None:
    print(f"  下载: {url}")
    try:
        r = requests.get(url, timeout=60)
        if r.status_code != 200:
            print(f"  跳过（HTTP {r.status_code}）")
            return None
        with zipfile.ZipFile(io.BytesIO(r.content)) as z:
            fname = z.namelist()[0]
            with z.open(fname) as f:
                return pd.read_csv(f, low_memory=False)
    except Exception as e:
        print(f"  失败: {e}")
        return None


def extract_net(df: pd.DataFrame, patterns: list[str]) -> pd.Series:
    """从原始 COT 表中提取管理资金净头寸时序。"""
    names = df["Market_and_Exchange_Names"].str.strip().str.upper()
    mask = pd.Series(False, index=df.index)
    for p in patterns:
        mask |= names.str.contains(p.upper(), regex=False)
    sub = df[mask].copy()
    if sub.empty:
        return pd.Series(dtype=float)

    sub["date"] = pd.to_datetime(sub["Report_Date_as_YYYY-MM-DD"], errors="coerce")
    sub = sub.dropna(subset=["date"])
    sub["net"] = (
        pd.to_numeric(sub["M_Money_Positions_Long_All"],  errors="coerce").fillna(0)
        - pd.to_numeric(sub["M_Money_Positions_Short_All"], errors="coerce").fillna(0)
    )
    s = sub.groupby("date")["net"].sum()
    return s.sort_index()


def main() -> None:
    all_frames: dict[str, list[pd.Series]] = {k: [] for k in CONTRACT_PATTERNS}

    # 历史合并文件（2006-2016）
    print("=== 下载历史文件 2006-2016 ===")
    df_hist = fetch_zip(URL_HIST)
    if df_hist is not None:
        for key, patterns in CONTRACT_PATTERNS.items():
            s = extract_net(df_hist, patterns)
            if not s.empty:
                all_frames[key].append(s)
                print(f"  {key}: {len(s)} 条")

    # 逐年下载 2017-2026
    import datetime
    current_year = datetime.date.today().year
    for year in range(2017, current_year + 1):
        print(f"=== 下载 {year} ===")
        df_yr = fetch_zip(URL_TEMPLATE.format(year=year))
        if df_yr is None:
            continue
        for key, patterns in CONTRACT_PATTERNS.items():
            s = extract_net(df_yr, patterns)
            if not s.empty:
                all_frames[key].append(s)
                print(f"  {key}: {len(s)} 条")

    # 合并并写入 Excel
    print(f"\n写入 {MASTER_FILE} ...")
    if not MASTER_FILE.exists():
        print(f"文件不存在: {MASTER_FILE}")
        return

    from openpyxl import load_workbook
    wb = load_workbook(MASTER_FILE)

    for key, frames in all_frames.items():
        if not frames:
            print(f"  {key}: 无数据，跳过")
            continue
        combined = pd.concat(frames).sort_index()
        combined = combined[~combined.index.duplicated(keep="last")]

        sheet_name = COT_SHEET_NAMES[key]
        desc = COT_DESCRIPTIONS[key]

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        ws.append(["日期", desc])
        for dt, val in combined.items():
            ws.append([dt.strftime("%Y-%m-%d"), float(val)])

        print(f"  {sheet_name}: {len(combined)} 条，{combined.index[0].date()} ~ {combined.index[-1].date()}")

    wb.save(MASTER_FILE)
    print("完成。")


if __name__ == "__main__":
    main()
