# -*- coding: utf-8 -*-
"""
根据「期货价格」子表计算并填充：
  日收益率、月收益率、近三月收益率、近六月收益率
"""

import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict

path = "数据.xlsx"

# 列对应：0日期-黄金, 1价格 | 2日期-银, 3价格 | 4日期-铜, 5价格 | 6日期-原油, 7价格
COLS = [(0, 1), (2, 3), (4, 5), (6, 7)]
NAMES = ["黄金", "银", "铜", "布伦特原油"]


def to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    return v


def build_series(rows, date_col, price_col):
    """从源表取一列日期、一列价格，得到按日期排序的 (date, price) 列表"""
    data = []
    for i in range(2, len(rows)):
        row = rows[i]
        if row is None:
            continue
        d = row[date_col] if date_col < len(row) else None
        p = row[price_col] if price_col < len(row) else None
        if d is None or p is None:
            continue
        if not isinstance(p, (int, float)):
            continue
        dt = to_date(d)
        if dt is None:
            continue
        data.append((dt, float(p)))
    data.sort(key=lambda x: x[0])
    return data


def daily_returns(series):
    """日收益率: (P_t - P_{t-1}) / P_{t-1}，返回 [(date, return), ...]"""
    out = []
    for i in range(1, len(series)):
        d, p = series[i]
        _, p_prev = series[i - 1]
        if p_prev and p_prev != 0:
            r = (p - p_prev) / p_prev
            out.append((d, r))
    return out


def month_end_prices(series):
    """按月末取价：每个 (year, month) 取该月最后一个交易日的 (date, price)"""
    by_month = defaultdict(list)  # (y, m) -> [(d, p), ...]
    for d, p in series:
        by_month[(d.year, d.month)].append((d, p))
    result = []
    for (y, m), vals in sorted(by_month.items()):
        last = max(vals, key=lambda x: x[0])
        result.append(last)
    return result


def monthly_returns(series):
    """月收益率: 月末价相比上月月末价"""
    ends = month_end_prices(series)
    out = []
    for i in range(1, len(ends)):
        (d, p), (_, p_prev) = ends[i], ends[i - 1]
        if p_prev and p_prev != 0:
            r = (p - p_prev) / p_prev
            out.append((d, r))
    return out


def quarterly_returns(series):
    """每三个月收益率：按自然季度（1-3月、4-6月、7-9月、10-12月），季度末价相对季度初价"""
    by_quarter = defaultdict(list)
    for d, p in series:
        q = (d.month - 1) // 3 + 1
        by_quarter[(d.year, q)].append((d, p))
    out = []
    for (y, q), vals in sorted(by_quarter.items()):
        vals_sorted = sorted(vals, key=lambda x: x[0])
        first_p = vals_sorted[0][1]
        last_d, last_p = vals_sorted[-1][0], vals_sorted[-1][1]
        if first_p and first_p != 0:
            out.append((last_d, (last_p - first_p) / first_p))
    return out


def semi_annual_returns(series):
    """每半年收益率：按上下半年（1-6月、7-12月），半年末价相对半年初价"""
    by_half = defaultdict(list)
    for d, p in series:
        h = 1 if d.month <= 6 else 2
        by_half[(d.year, h)].append((d, p))
    out = []
    for (y, h), vals in sorted(by_half.items()):
        vals_sorted = sorted(vals, key=lambda x: x[0])
        first_p = vals_sorted[0][1]
        last_d, last_p = vals_sorted[-1][0], vals_sorted[-1][1]
        if first_p and first_p != 0:
            out.append((last_d, (last_p - first_p) / first_p))
    return out


def align_to_common_dates(series_list, fill_na=None):
    """
    将多个 [(date, value), ...] 对齐到所有日期的并集，返回
    [(date, [v_gold, v_silver, v_copper, v_oil]), ...]
    """
    all_dates = set()
    for s in series_list:
        for d, _ in s:
            all_dates.add(d)
    all_dates = sorted(all_dates)
    by_date = [dict(s) for s in series_list]
    out = []
    for d in all_dates:
        row = [d]
        for i, m in enumerate(by_date):
            row.append(m.get(d, fill_na))
        out.append(row)
    return out


def main():
    print("正在读取 期货价格 表...")
    wb_read = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws_src = wb_read["期货价格"]
    rows = list(ws_src.iter_rows(values_only=True))
    wb_read.close()

    # 四个品种的 (date, price) 序列
    series_list = []
    for date_col, price_col in COLS:
        s = build_series(rows, date_col, price_col)
        series_list.append(s)
        print(f"  {NAMES[series_list.index(s)]}: {len(s)} 条")

    # 计算四种收益率（每个品种一条序列）
    # 近三月 = 每三个月收益率（按季度）；近六月 = 每半年收益率
    daily_list = [daily_returns(s) for s in series_list]
    monthly_list = [monthly_returns(s) for s in series_list]
    m3_list = [quarterly_returns(s) for s in series_list]
    m6_list = [semi_annual_returns(s) for s in series_list]

    # 对齐到公共日期
    daily_aligned = align_to_common_dates(daily_list, fill_na=None)
    monthly_aligned = align_to_common_dates(monthly_list, fill_na=None)
    m3_aligned = align_to_common_dates(m3_list, fill_na=None)
    m6_aligned = align_to_common_dates(m6_list, fill_na=None)

    print("正在写入四个收益率子表...")
    wb = openpyxl.load_workbook(path)

    def write_return_sheet(wb, sheet_name, headers, data_rows):
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        for c, h in enumerate(headers, 1):
            ws.cell(1, c, h)
        for r, row in enumerate(data_rows, 2):
            for c, val in enumerate(row, 1):
                ws.cell(r, c, val)
        return ws

    headers = ["日期"] + [f"{n}收益率" for n in NAMES]

    write_return_sheet(wb, "日收益率", headers, daily_aligned)
    write_return_sheet(wb, "月收益率", headers, monthly_aligned)
    write_return_sheet(wb, "近三月收益率", headers, m3_aligned)
    write_return_sheet(wb, "近六月收益率", headers, m6_aligned)

    wb.save(path)
    print("已保存。四个子表：日收益率、月收益率、近三月收益率、近六月收益率 已根据期货价格计算并填好。")


if __name__ == "__main__":
    main()
