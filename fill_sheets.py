# -*- coding: utf-8 -*-
"""根据「期货价格」子表数据，创建并填充四个品种子表：黄金、银、铜、布伦特原油"""

import openpyxl

path = "数据.xlsx"

# 第一步：只读方式读取期货价格表的数据（保留原表公式不变）
print("正在读取 期货价格 表数据...")
wb_read = openpyxl.load_workbook(path, read_only=True, data_only=True)
ws_src = wb_read["期货价格"]
rows = list(ws_src.iter_rows(values_only=True))
wb_read.close()

# 列对应：0日期-黄金, 1价格 | 2日期-银, 3价格 | 4日期-铜, 5价格 | 6日期-原油, 7价格
# 数据从第3行开始（索引2），前两行为表头
header_row0 = rows[0]  # 中文表头
header_row1 = rows[1]  # 代码行

def build_sheet_data(rows, date_col, price_col, name_cn):
    """从源表取一列日期、一列价格，去掉空行"""
    data = []
    for i in range(2, len(rows)):
        row = rows[i]
        if row is None:
            continue
        d = row[date_col] if date_col < len(row) else None
        p = row[price_col] if price_col < len(row) else None
        if d is None and p is None:
            continue
        data.append((d, p))
    return data

gold_data = build_sheet_data(rows, 0, 1, "黄金")
silver_data = build_sheet_data(rows, 2, 3, "银")
copper_data = build_sheet_data(rows, 4, 5, "铜")
oil_data = build_sheet_data(rows, 6, 7, "布伦特原油")

print(f"黄金: {len(gold_data)} 行")
print(f"银: {len(silver_data)} 行")
print(f"铜: {len(copper_data)} 行")
print(f"布伦特原油: {len(oil_data)} 行")

# 第二步：以可写方式打开工作簿（不破坏期货价格表里的公式）
print("正在写入四个子表...")
wb = openpyxl.load_workbook(path)

def write_sheet(wb, sheet_name, title_date, title_price, data):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.cell(1, 1, title_date)
    ws.cell(1, 2, title_price)
    for i, (d, p) in enumerate(data, start=2):
        ws.cell(i, 1, d)
        ws.cell(i, 2, p)
    return ws

write_sheet(wb, "黄金", "日期", "期货收盘价(连续):COMEX黄金", gold_data)
write_sheet(wb, "银", "日期", "期货收盘价(连续):COMEX银", silver_data)
write_sheet(wb, "铜", "日期", "期货收盘价(活跃合约):COMEX铜", copper_data)
write_sheet(wb, "布伦特原油", "日期", "期货收盘价(活跃合约):MICEX 布伦特原油", oil_data)

wb.save(path)
print("已保存 数据.xlsx，四个子表已根据期货价格表填写完成。")
