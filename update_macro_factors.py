# -*- coding: utf-8 -*-
"""
更新宏观因子数据（降序存储 sheet 的追加补充）：
  - 美国ISM制造业PMI  → FRED NAPM
  - 美国PPI           → FRED PPIACO
  - 美国CPI           → FRED CPIAUCSL
  - LME铜库存         → akshare
  - COMEX黄金/白银持仓量 → akshare
"""

import subprocess
import io
import requests
import pandas as pd
import openpyxl
from io import StringIO

MASTER_FILE = "大宗商品轮动_数据2.xlsx"


# ── FRED 下载 ────────────────────────────────────────────────
def download_fred(series_id: str, col_name: str) -> pd.DataFrame | None:
    url = f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}"
    try:
        # curl with longer timeout
        result = subprocess.run(["curl", "-s", "--max-time", "60", "--retry", "3", url],
                                capture_output=True, text=True, timeout=90)
        text = result.stdout if result.returncode == 0 and result.stdout.startswith("observation") else None
        if not text:
            resp = requests.get(url, timeout=60, headers={"User-Agent": "Mozilla/5.0"})
            resp.raise_for_status()
            text = resp.text
        df = pd.read_csv(StringIO(text))
        df.columns = ["日期", col_name]
        df["日期"] = pd.to_datetime(df["日期"])
        df[col_name] = pd.to_numeric(df[col_name], errors="coerce")
        return df.dropna().sort_values("日期", ascending=False).reset_index(drop=True)
    except Exception as e:
        print(f"    ❌ FRED {series_id}: {e}")
        return None


# ── 获取 sheet 当前最新日期（降序存储，第一行数据为最新）────────
def get_sheet_latest(wb, sheet_name) -> pd.Timestamp | None:
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=2, max_row=3, values_only=True))
    for r in rows:
        if r and r[0] is not None:
            try:
                return pd.Timestamp(r[0])
            except Exception:
                pass
    return None


# ── 将新行插入 sheet 最前（降序，第2行起） ───────────────────
def prepend_rows(ws, new_df: pd.DataFrame):
    """
    new_df: columns=[日期, 值]，按降序排列（新日期在前）
    插入到 header 后第2行，现有数据下移。
    """
    if new_df.empty:
        return 0

    # 读取现有数据（行2起）
    existing = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            existing.append(list(row))

    # 合并：新数据在前，旧数据在后
    header = list(new_df.columns)
    new_rows = [[r["日期"].date() if hasattr(r["日期"], "date") else r["日期"],
                 r.iloc[1]] for _, r in new_df.iterrows()]
    all_rows = new_rows + existing

    # 清空并重写
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    for i, row in enumerate(all_rows, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    return len(new_rows)


# ── 主流程 ───────────────────────────────────────────────────
def main():
    wb = openpyxl.load_workbook(MASTER_FILE)

    # ── 1. 美国ISM制造业PMI（FRED NAPM，月度）──
    print("\n【美国ISM制造业PMI】")
    latest = get_sheet_latest(wb, "美国ISM制造业PMI")
    print(f"  当前最新: {latest.date() if latest else '无'}")
    df = download_fred("NAPM", "美国ISM制造业PMI")
    if df is not None:
        new = df[df["日期"] > latest] if latest else df
        if not new.empty:
            ws = wb["美国ISM制造业PMI"]
            n = prepend_rows(ws, new)
            print(f"  ✅ 追加 {n} 行，最新: {new['日期'].iloc[0].date()}")
        else:
            print("  ✅ 已是最新")

    # ── 2. 美国PPI（FRED PPIACO）──
    print("\n【美国PPI】")
    latest = get_sheet_latest(wb, "美国PPI")
    print(f"  当前最新: {latest.date() if latest else '无'}")
    df = download_fred("PPIACO", "美国PPI(生产者价格指数)")
    if df is not None:
        new = df[df["日期"] > latest] if latest else df
        if not new.empty:
            ws = wb["美国PPI"]
            n = prepend_rows(ws, new)
            print(f"  ✅ 追加 {n} 行，最新: {new['日期'].iloc[0].date()}")
        else:
            print("  ✅ 已是最新")

    # ── 3. 美国CPI（FRED CPIAUCSL）──
    print("\n【美国CPI】")
    latest = get_sheet_latest(wb, "美国CPI")
    print(f"  当前最新: {latest.date() if latest else '无'}")
    df = download_fred("CPIAUCSL", "美国CPI(消费者价格指数)")
    if df is not None:
        new = df[df["日期"] > latest] if latest else df
        if not new.empty:
            ws = wb["美国CPI"]
            n = prepend_rows(ws, new)
            print(f"  ✅ 追加 {n} 行，最新: {new['日期'].iloc[0].date()}")
        else:
            print("  ✅ 已是最新")

    # ── 4. LME铜库存（akshare macro_euro_lme_stock）──
    print("\n【LME铜库存】")
    try:
        import akshare as ak
        ws_lme = wb["LME铜库存"]
        # LME铜库存是降序存储（最新在前），找第一行数据的日期
        lme_rows = list(ws_lme.iter_rows(min_row=2, max_row=3, values_only=True))
        lme_latest = None
        for r in lme_rows:
            if r and r[0] is not None:
                try:
                    lme_latest = pd.Timestamp(r[0])
                    break
                except Exception:
                    pass
        print(f"  当前最新: {lme_latest.date() if lme_latest else '无'}")

        df_lme = ak.macro_euro_lme_stock()
        df_lme = df_lme[["日期", "铜-库存"]].copy()
        df_lme.columns = ["日期", "LME铜库存(吨)"]
        df_lme["日期"] = pd.to_datetime(df_lme["日期"])
        df_lme = df_lme.dropna().sort_values("日期", ascending=False)
        print(f"  akshare最新: {df_lme['日期'].iloc[0].date()}")

        new = df_lme[df_lme["日期"] > lme_latest] if lme_latest else df_lme
        if not new.empty:
            n = prepend_rows(ws_lme, new.reset_index(drop=True))
            print(f"  ✅ 追加 {n} 行，最新: {new['日期'].iloc[0].date()}")
        else:
            print("  ✅ 已是最新")
    except Exception as e:
        print(f"  ❌ LME铜库存失败: {e}")

    # ── 5. COMEX黄金持仓量（akshare futures_comex_inventory，升序存储）──
    print("\n【COMEX黄金持仓量】")
    try:
        import akshare as ak
        ws_oi = wb["comex黄金持仓量"]
        rows_oi = list(ws_oi.iter_rows(values_only=True))
        last_dt = None
        for r in rows_oi[1:]:
            if r and r[0] is not None:
                try:
                    last_dt = pd.Timestamp(r[0])
                except Exception:
                    pass
        print(f"  当前最新: {last_dt.date() if last_dt else '无'}")

        df_comex = ak.futures_comex_inventory()
        # 返回: 序号, 日期, COMEX黄金库存量-吨, COMEX黄金库存量-盎司
        df_comex = df_comex[["日期", "COMEX黄金库存量-吨"]].copy()
        df_comex.columns = ["日期", "COMEX黄金库存量(吨)"]
        df_comex["日期"] = pd.to_datetime(df_comex["日期"])
        df_comex = df_comex.dropna().sort_values("日期")
        print(f"  akshare最新: {df_comex['日期'].iloc[-1].date()}")

        new = df_comex[df_comex["日期"] > last_dt] if last_dt else df_comex
        if not new.empty:
            last_row = max((i+1 for i, r in enumerate(
                ws_oi.iter_rows(values_only=True)) if any(v is not None for v in r)), default=1)
            for _, row in new.iterrows():
                last_row += 1
                ws_oi.cell(row=last_row, column=1, value=row["日期"].date())
                ws_oi.cell(row=last_row, column=2, value=float(row.iloc[1]))
            print(f"  ✅ 追加 {len(new)} 行，最新: {new['日期'].iloc[-1].date()}")
        else:
            print("  ✅ 已是最新")
        print("  ℹ️  comex白银持仓量 akshare 无对应接口，需从 Wind 手动更新")
    except Exception as e:
        print(f"  ❌ COMEX黄金持仓量失败: {e}")

    wb.save(MASTER_FILE)
    print("\n✅ 保存完成。")


if __name__ == "__main__":
    main()
